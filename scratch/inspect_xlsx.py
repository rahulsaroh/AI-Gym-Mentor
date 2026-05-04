import openpyxl
import json

def inspect_xlsx(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    report = {}
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        # headers are in the first row
        headers = []
        for cell in sheet[1]:
            headers.append(str(cell.value) if cell.value is not None else "")
        
        rows = []
        # Get first 5 rows of data
        count = 0
        for row in sheet.iter_rows(min_row=2, max_row=6, values_only=True):
            if any(row):
                rows.append([str(v) if v is not None else "" for v in row])
                count += 1
        
        # approximate total rows by iterating (max_row is sometimes unreliable or missing in some versions/files)
        total_rows = 0
        for row in sheet.iter_rows(values_only=True):
            if any(row):
                total_rows += 1

        report[sheet_name] = {
            "headers": headers,
            "sample_rows": rows,
            "total_rows": total_rows
        }
    return report

if __name__ == "__main__":
    file1 = r"D:\Apps\Completed\AI Gym Buddy\AI Gym Mentor\gym_report_220426.xlsx"
    file2 = r"D:\Apps\Completed\AI Gym Buddy\AI Gym Mentor\Copy of gym_report_22042026.xlsx"
    
    print("Inspecting file 1...")
    try:
        report1 = inspect_xlsx(file1)
        with open("xlsx_report1.json", "w") as f:
            json.dump(report1, f, indent=2)
    except Exception as e:
        import traceback
        print(f"Error inspecting file 1: {e}")
        traceback.print_exc()

    print("Inspecting file 2...")
    try:
        report2 = inspect_xlsx(file2)
        with open("xlsx_report2.json", "w") as f:
            json.dump(report2, f, indent=2)
    except Exception as e:
        import traceback
        print(f"Error inspecting file 2: {e}")
        traceback.print_exc()
